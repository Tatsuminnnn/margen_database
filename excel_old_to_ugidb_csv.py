"""
胃癌台帳 1990-2019 Excel → UGI_DB CSV 変換スクリプト
=====================================================
旧データベース (167列, 13版/14版混在) を csv_import.py 互換CSVに変換。

変換方針:
  - 14版カラムを優先 → 空なら旧規約カラムからfallback
  - 15版コードブックにマッピング
  - 存在しない項目(ASA, PS, HER2等) はNULLのまま

使い方:
    python excel_old_to_ugidb_csv.py input.xlsx output.csv [--dry-run]
"""

import sys
import os
import re
import csv
import io
from datetime import datetime, date
from collections import OrderedDict

try:
    import openpyxl
except ImportError:
    print("openpyxl が必要です: pip install openpyxl")
    sys.exit(1)


# ============================================================
# 日付ヘルパー
# ============================================================
def _fmt_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y年%m月%d日"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _safe_int(val):
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _safe_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _cell(ws, row, col):
    v = ws.cell(row=row, column=col).value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v in ("", "該当なし", "-", "−", "ー", "不明", "記載なし"):
            return None
    return v


def _map_val(val, mapping, default=None):
    if val is None:
        return default
    s = str(val).strip()
    if s in mapping:
        return mapping[s]
    for key, code in mapping.items():
        if key in s:
            return code
    return default


# ============================================================
# 性別  M/F → 1/2
# ============================================================
SEX_MAP = {"M": 1, "F": 2, "男": 1, "女": 2}

# ============================================================
# 疾患分類
# ============================================================
DISEASE_CLASS_MAP = {
    "初発胃癌": 1,
    "初発胃癌（ESD後）": 2, "初発胃癌（EMR後）": 2,
    "残胃癌": 3, "残胃の癌": 3,
    "GIST": 4, "gist": 4,
    "悪性リンパ腫": 5,  # B cell assumed
    "その他の悪性腫瘍": 8,
    "その他": 8,
    "その他の腫瘍": 8,
    "再発": 1,  # 再発も元は胃癌 → 初発胃癌として登録、再発フラグ別途
}

# ============================================================
# 肉眼型 (旧 → 15版コード)
# ============================================================
MACRO_TYPE_MAP = {
    "0型": 0, "1型": 1, "2型": 2, "3型": 3, "4型": 4, "5型": 5,
}

TYPE0_MAP = {
    "Ⅰ型": 1, "I型": 1, "Ⅱa型": 2, "IIa型": 2, "Ⅱb型": 3,
    "Ⅱc型": 4, "IIc型": 4, "Ⅲ型": 5, "III型": 5,
    "Ⅱa": 2, "Ⅱb": 3, "Ⅱc": 4,
    "IIa": 2, "IIb": 3, "IIc": 4,
}

# ============================================================
# 組織型
# ============================================================
HISTOLOGY_MAP = {
    "tub1": 1, "tub2": 2, "por1": 3, "por2": 4, "por": 3,
    "sig": 5, "muc": 6, "pap": 7,
    "SCM": 8,  # neuroendocrine carcinoma ≈ SCM
    "ASQ": 9,  # adenosquamous
    "ub2": 2,  # typo for tub2
    "adeno": 7,  # adenocarcinoma NOS → pap as closest
    "OTH": None, "TUM": None, "CND": None, "MIS": None,
}

# ============================================================
# INF (旧: INF-α/β/γ → 新: 1/2/3)
# ============================================================
INF_MAP = {
    "INF-α": 1, "INF-β": 2, "INF-γ": 3,
    "INFa": 1, "INFb": 2, "INFc": 3, "INFα": 1, "INFβ": 2, "INFγ": 3,
}

# ============================================================
# ly/v (旧 ly0-3 → 新 ly0-3; ly1a/b/c → 1/2/3)
# ============================================================
LY_MAP = {
    "ly0": 0, "ly1": 1, "ly2": 2, "ly3": 3,
    "ly1a": 1, "ly1b": 2, "ly1c": 3,
    "Ly0": 0, "Ly1": 1, "Ly1a": 1, "Ly1b": 2,
}
V_MAP = {
    "v0": 0, "v1": 1, "v2": 2, "v3": 3,
    "v1a": 1, "v1b": 2, "v1c": 3,
    "V0": 0, "V1": 1, "V1a": 1, "V1b": 2,
}

# ============================================================
# cT 変換  (14版優先 → 旧規約fallback → 15版コード)
# ============================================================
CT_14_MAP = {
    "cT1a(M)": 1, "cT1b1(SM1)": 2, "cT1b2(SM2)": 2, "cT1b": 2,
    "cT2(MP)": 3, "cT2": 3,
    "cT3(SS)": 4,
    "cT4a(SE)": 5, "cT4a": 5,
    "cT4b(SI)": 6, "cT4b": 6,
    "cTX": 9,
}

# 旧規約 cT (13版以前: cT1-4)
CT_OLD_MAP = {
    "cT1": 2,  # cT1 → SM相当 (M/SM未区別 → SM扱い)
    "cT2": 3,  # MP
    "cT3": 4,  # SS
    "cT4": 5,  # SE (SI区別なし → SE扱い)
    "cTX": 9,
}

# 旧 深達度テキスト
CT_TEXT_MAP = {
    "M": 1, "SM": 2, "MP": 3, "SS": 4, "SE": 5, "SI": 6,
}

def _parse_ct(val_14, val_old, val_text):
    """14版 → 旧cT → テキストの順でfallback。"""
    if val_14:
        r = _map_val(val_14, CT_14_MAP)
        if r is not None:
            return r
    if val_old:
        r = _map_val(val_old, CT_OLD_MAP)
        if r is not None:
            return r
    if val_text:
        r = _map_val(val_text, CT_TEXT_MAP)
        if r is not None:
            return r
    return None


# ============================================================
# pT 変換
# ============================================================
PT_14_MAP = {
    "pT0": 0,
    "pT1a(M)": 1, "pT1b1(SM1)": 2, "pT1b2(SM2)": 2, "pT1b": 2,
    "pT2(MP)": 3, "pT2": 3, "T2": 3,
    "pT3(SS)": 4,
    "pT4a(SE)": 5, "pT4a": 5,
    "pT4b(SI)": 6, "pT4b": 6,
    "pTX": 9,
}

PT_OLD_MAP = {
    "pT0": 0, "pT1": 2, "pT2": 3, "pT3": 4,
    "pT4": 5, "pT4a": 5, "pT4b": 6, "pTX": 9,
}

PT_TEXT_MAP = {
    "pM": 1, "pSM": 2, "pMP": 3, "pSS": 4, "pSE": 5, "pSI": 6,
    "T0": 0,
}

def _parse_pt(val_14, val_old, val_text):
    if val_14:
        r = _map_val(val_14, PT_14_MAP)
        if r is not None:
            return r
    if val_old:
        r = _map_val(val_old, PT_OLD_MAP)
        if r is not None:
            return r
    if val_text:
        r = _map_val(val_text, PT_TEXT_MAP)
        if r is not None:
            return r
    return None


# ============================================================
# cN / pN
# ============================================================
# 14版 cN (15版互換)
CN_14_MAP = {
    "cN0": 0, "cN1": 1, "cN2": 2, "cN3a": 3, "cN3b": 4,
    "ｃN+": 1, "cN+": 1, "cNX": 9,
}
# 旧 cN
CN_OLD_MAP = {
    "cN0": 0, "cN1": 1, "cN2": 2, "cN3": 3, "cNX": 9,
}

def _parse_cn(val_14, val_old):
    if val_14:
        r = _map_val(val_14, CN_14_MAP)
        if r is not None:
            return r
    if val_old:
        r = _map_val(val_old, CN_OLD_MAP)
        if r is not None:
            return r
    return None

# 14版 pN
PN_14_MAP = {
    "pN0": 0, "pN1": 1, "pN2": 2,
    "pN3a": 3, "pN3b": 4, "pN3ｃ": 4,
    "pNX": 9,
}
# 旧 pN
PN_OLD_MAP = {
    "pN0": 0, "pN1": 1, "pN2": 2,
    "pN3": 3, "pN3a": 3, "pN3b": 4, "pN4": 4,
    "pNX": 9,
}

def _parse_pn(val_14, val_old):
    if val_14:
        r = _map_val(val_14, PN_14_MAP)
        if r is not None:
            return r
    if val_old:
        r = _map_val(val_old, PN_OLD_MAP)
        if r is not None:
            return r
    return None


def _parse_pn_from_count(count_val, fallback_col78=None):
    """col143（全転移陽性個数）→ 14/15版 pN コード。

    変換表:
      9999 → 9 (NX)
      0    → 0 (N0)
      1-2  → 1 (N1)
      3-6  → 2 (N2)
      7-15 → 3 (N3a)
      16+  → 4 (N3b)
    col143 が欠損の場合は col78（14版pN テキスト）にfallback。
    """
    if count_val is not None:
        try:
            n = int(float(count_val))
            if n == 9999:
                return 9   # NX
            if n == 0:
                return 0   # N0
            elif n <= 2:
                return 1   # N1
            elif n <= 6:
                return 2   # N2
            elif n <= 15:
                return 3   # N3a
            else:
                return 4   # N3b
        except (ValueError, TypeError):
            pass
    # fallback: col78（14版pN）テキスト変換
    return _parse_pn(fallback_col78, None)


# ============================================================
# cH, cP, cM, pH, pP, pM
# ============================================================
CH_MAP = {"cH0": 0, "cH1": 1}
CP_MAP = {"cP0": 0, "cP1": 1, "cPX": None}
CM_MAP = {"cM0": 0, "cM1": 1, "cMX": 9}
PH_MAP = {"pH0": 0, "pH1": 1, "pHX": None}
PP_MAP = {"pP0": 0, "pP1": 1, "pPX": None}
PM_MAP = {"pM0": 0, "pM1": 1, "pMX": 9}

# ============================================================
# CY
# ============================================================
CY_MAP = {"pCY0": 0, "pCY1": 1, "pCYX": 9, "判定不能": 9}

# ============================================================
# 断端 pPM / pDM
# ============================================================
def _parse_margin(val):
    if val is None:
        return None
    s = str(val).strip()
    if "(+)" in s or "陽性" in s or "PM1" in s or "DM1" in s:
        return 1
    if "(-)" in s or "陰性" in s or "PM0" in s or "DM0" in s:
        return 0
    return None

# ============================================================
# 遺残 pR
# ============================================================
RESIDUAL_MAP = {"pR0": 0, "pR1": 1, "pR2": 2, "pRX": 8}

# ============================================================
# Stage (14版→15版マッピング) — 全角ローマ数字対応
# ============================================================
def _normalize_stage(s):
    """全角ローマ数字→半角変換。"""
    if s is None:
        return None
    s = str(s).strip()
    s = s.replace("Ⅰ", "I").replace("Ⅱ", "II").replace("Ⅲ", "III").replace("Ⅳ", "IV")
    # remove leading p/c/f
    s = re.sub(r"^[pcfｐｃｆyｙ]+", "", s)
    return s

C_STAGE_MAP = {
    "IA": 1, "IB": 1,  # 15版ではIAのみ → cStage=1 (I)
    "I": 1,
    "IIA": 2, "IIB": 3, "II": 2,
    "IIIA": 4, "IIIB": 4, "IIIC": 4, "III": 4,
    "IVA": 5, "IVB": 6, "IV": 5,
}

P_STAGE_MAP = {
    "0": 0,
    "IA": 1, "IB": 2, "I": 1,
    "IIA": 3, "IIB": 4, "II": 3,
    "IIIA": 5, "IIIB": 6, "IIIC": 7, "III": 5,
    "IV": 8, "IVA": 8, "IVB": 8,
}

def _parse_stage(val_15, val_14, val_old, stage_map):
    """15版 → 14版 → 旧の順でfallback。"""
    for v in [val_15, val_14, val_old]:
        ns = _normalize_stage(v)
        if ns and ns in stage_map:
            return stage_map[ns]
    return None


# ============================================================
# アプローチ
# ============================================================
APPROACH_MAP = {
    "開腹": 1,
    "腹腔鏡・腹腔鏡補助": 2, "腹腔鏡": 2, "腹腔鏡補助": 7,
    "ロボット支援下": 3, "ロボット支援": 3,
    "開胸": 4, "胸腔鏡": 5, "内視鏡": 9,
    "横隔膜切開": 1,  # 開腹相当
    "開胸開腹(連続)": 4, "開胸開腹（連続）": 4,
    "開胸開腹（非連続）": 4, "開胸開腹(非連続)": 4,
}

# ============================================================
# 術式
# ============================================================
PROCEDURE_MAP = {
    "幽門側胃切除": 1, "DG": 1,
    "胃全摘": 2, "TG": 2,
    "噴門側胃切除": 3, "PG": 3,
    "胃局所切除": 4, "局所切除": 4,
    "PPG": 5, "PPNTG": 5, "幽門保存胃切除": 5,
    "残胃全摘": 6,
    "審査腹腔鏡": 7, "試験開腹": 7,
    "吻合": 8, "バイパス手術": 8, "バイパス": 8,
    "胃腸ろう": 9, "胃粘膜切除": 9, "胃瘻造設": 9,
}

# ============================================================
# 郭清 (D1+α, D1+β → D1+ (=2) に統合)
# ============================================================
def _parse_dissection(val_new, val_old):
    """新旧郭清コードの統合変換。"""
    DISS_NEW = {"D0": 0, "D1": 1, "D1+": 2, "D2": 3, "D2+": 4, "D3": 5}
    DISS_OLD = {
        "D0": 0, "D1": 1,
        "D1+α": 2, "D1+β": 2, "D1+": 2,
        "D2": 3, "D3": 5,
        "なし": 0,
    }
    if val_new:
        r = _map_val(val_new, DISS_NEW)
        if r is not None:
            return r
    if val_old:
        r = _map_val(val_old, DISS_OLD)
        if r is not None:
            return r
    return None


# ============================================================
# 再建法
# ============================================================
RECON_MAP = {
    "Billroth-Ⅰ法": 1, "Billroth-I法": 1, "BI法": 1, "B-I": 1,
    "Billroth-Ⅱ法": 2, "Billroth-II法": 2, "BII法": 2, "B-II": 2,
    "Roux-Y": 3, "Roux-en-Y法": 3, "Roux-en-Y": 3, "R-Y": 3, "RY": 3,
    "食道残胃吻合": 4,
    "ｄouble tract": 5, "double tract": 5, "ダブルトラクト": 5,
    "空腸間置": 6,
    "幽門保存": 7,  # 再建なし相当 (PPGは消化管連続)
    "非切除": 7,  # 非切除 → 再建なし
    "その他の再建": 9,
}

# ============================================================
# 合併切除
# ============================================================
COMBINED_KEYWORDS = {
    "comb_splenectomy": ["脾", "spleen"],
    "comb_cholecystectomy": ["胆嚢", "胆摘"],
    "comb_distal_pancreatectomy": ["膵", "pancreas"],
    "comb_transverse_colectomy": ["結腸", "横行結腸"],
    "comb_partial_hepatectomy": ["肝"],
    "comb_diaphragm": ["横隔膜"],
    "comb_ovary": ["卵巣"],
}

def _decompose_combined(val1, val2):
    flags = {k: 0 for k in COMBINED_KEYWORDS}
    combined = " ".join(str(v) for v in [val1, val2]
                        if v and "なし" not in str(v))
    if not combined.strip():
        return flags
    for col, keywords in COMBINED_KEYWORDS.items():
        for kw in keywords:
            if kw in combined:
                flags[col] = 1
                break
    return flags


# ============================================================
# 合併症
# ============================================================
COMP_KEYWORDS = {
    "comp_anastomotic_leak": ["縫合不全"],
    "comp_pancreatic_fistula": ["膵液漏", "膵瘻"],
    "comp_ssi": ["感染", "SSI", "創感染"],
    "comp_anastomotic_stricture": ["狭窄"],
    "comp_pneumonia": ["呼吸器", "肺炎"],
    "comp_ileus": ["イレウス", "腸閉塞"],
    "comp_intra_abd_abscess": ["腹腔内膿瘍", "膿瘍"],
    "comp_bleeding": ["出血"],
    "comp_perforation": ["穿孔"],
    "comp_dge": ["DGE", "胃排泄遅延"],
    "comp_cholelithiasis": ["胆石"],
    "comp_dic": ["DIC"],
    "comp_hepatic_failure": ["肝疾患", "肝不全"],
    "comp_cardiac": ["循環器", "心"],
    "comp_renal_failure": ["腎疾患", "腎不全"],
}

def _decompose_complications(val1, val2):
    flags = {k: 0 for k in COMP_KEYWORDS}
    combined = " ".join(str(v) for v in [val1, val2]
                        if v and str(v).strip() not in ("なし", "不明"))
    if not combined.strip():
        flags["op_complication_yn"] = 0
        return flags
    flags["op_complication_yn"] = 1
    for col, keywords in COMP_KEYWORDS.items():
        for kw in keywords:
            if kw in combined:
                flags[col] = 1
                break
    return flags


# ============================================================
# 縫合不全 → CD grade (旧DBはminor/majorの区別のみ)
# ============================================================
def _estimate_cd_grade(comp1, comp2):
    """旧DBのcomp文字列からCD gradeを推定。限界あり。"""
    combined = " ".join(str(v) for v in [comp1, comp2]
                        if v and str(v).strip() not in ("なし", "不明"))
    if not combined.strip() or combined.strip() == "その他":
        return 0  # Grade 0
    if "(major)" in combined or "major" in combined.lower():
        return 3  # Grade IIIa 推定
    if "(minor)" in combined or "minor" in combined.lower():
        return 2  # Grade II 推定
    # 合併症ありだがgrade不明 → GradeII 仮設定
    return 2


# ============================================================
# 生死・死因 (1カラム) → vital_status + death_cause 分離
# ============================================================
def _parse_vital_death(val):
    """'生存中', '死亡', '腹膜再発（癌死）' 等を分離。"""
    vital_status = None
    death_cause = None
    if val is None:
        return None, None
    s = str(val).strip()

    if s == "生存中":
        return 1, None
    if s == "死亡":
        return 2, 1  # 癌死推定
    if s == "在院死":
        return 4, 6  # 手術関連死推定
    if "手術関連死" in s:
        return 4, 6

    if "癌死" in s:
        vital_status = 2  # 原病死
        if "腹膜" in s:
            death_cause = 2
        elif "肝" in s:
            death_cause = 3
        elif "肺" in s:
            death_cause = 4
        elif "リンパ節" in s:
            death_cause = 5
        elif "局所" in s or "残胃" in s:
            death_cause = 1
        elif "遠隔" in s:
            death_cause = 5
        elif "形式不明" in s:
            death_cause = 1
        else:
            death_cause = 1  # 原発巣増悪
        return vital_status, death_cause

    if "他病死" in s:
        return 3, 7
    if "他癌死" in s:
        return 3, 10  # 他病死(他癌)

    if "原因不明" in s or "不明" in s:
        return 9, None

    return 9, None  # 不明


# ============================================================
# 再発形式
# ============================================================
RECURRENCE_MAP_YN = {
    "再発": 1, "あり": 1, "なし": 0, "無再発": 0,
}

RECURRENCE_SITE_KEYWORDS = {
    "rec_peritoneal": ["腹膜"],
    "rec_liver": ["肝"],
    "rec_lung": ["肺"],
    "rec_lymph_node": ["リンパ節"],
    "rec_local": ["局所", "残胃", "断端"],
    "rec_bone": ["骨"],
    "rec_brain": ["脳"],
    "rec_ovary": ["卵巣"],
}

def _decompose_recurrence(val1, val2, val3):
    flags = {k: 0 for k in RECURRENCE_SITE_KEYWORDS}
    combined = " ".join(str(v) for v in [val1, val2, val3]
                        if v and str(v).strip() not in ("なし", "不明"))
    if not combined.strip():
        return flags
    for col, keywords in RECURRENCE_SITE_KEYWORDS.items():
        for kw in keywords:
            if kw in combined:
                flags[col] = 1
                break
    return flags


# ============================================================
# 併存疾患
# ============================================================
COMOR_KEYWORDS = {
    "comor_hypertension": ["高血圧", "HT"],
    "comor_cardiovascular": ["心疾患", "心不全", "不整脈", "弁膜"],
    "comor_cerebrovascular": ["脳梗塞", "脳出血", "脳卒中"],
    "comor_respiratory": ["呼吸器", "喘息", "COPD", "肺気腫"],
    "comor_renal": ["腎疾患", "腎不全", "CKD"],
    "comor_renal_dialysis": ["透析"],
    "comor_hepatic": ["肝疾患", "肝硬変", "肝障害"],
    "comor_diabetes": ["糖尿病", "DM"],
    "comor_endocrine": ["甲状腺", "内分泌"],
}

def _decompose_comorbidities(val1, val2):
    flags = {k: 0 for k in COMOR_KEYWORDS}
    combined = " ".join(str(v) for v in [val1, val2]
                        if v and "なし" not in str(v))
    if not combined.strip():
        return flags
    for col, keywords in COMOR_KEYWORDS.items():
        for kw in keywords:
            if kw in combined:
                flags[col] = 1
                break
    return flags


# ============================================================
# 症状
# ============================================================
SYMPTOM_KEYWORDS = {
    "sym_asymptomatic": ["なし", "無症状", "検診"],
    "sym_epigastric_pain": ["腹痛", "心窩部痛", "胃痛"],
    "sym_dysphagia": ["嚥下", "つかえ"],
    "sym_weight_loss": ["体重減少", "るいそう"],
    "sym_anemia": ["貧血"],
    "sym_melena": ["黒色便", "下血", "メレナ", "タール便"],
    "sym_hematemesis": ["吐血", "嘔吐"],
    "sym_nausea_vomiting": ["嘔気", "悪心"],
    "sym_abdominal_distension": ["腹部膨満", "腹満"],
    "sym_obstruction": ["閉塞", "通過障害"],
}

def _decompose_symptoms(val):
    flags = {k: 0 for k in SYMPTOM_KEYWORDS}
    if val is None:
        return flags
    s = str(val).strip()
    if not s or s in ("なし", "無症状"):
        flags["sym_asymptomatic"] = 1
        return flags
    if s == "その他":
        return flags  # 全てゼロ
    for col, keywords in SYMPTOM_KEYWORDS.items():
        for kw in keywords:
            if kw in s:
                flags[col] = 1
                break
    return flags


# ============================================================
# NAC regimen (テキスト → コード)
# ============================================================
NAC_REGIMEN_PATTERNS = [
    (r"SOX", 1), (r"SP$|S-?1\+CDDP", 2), (r"DOS|DTX", 3),
    (r"FLOT", 4), (r"Nivo.*SOX", 5), (r"Nivo.*CAPOX", 6),
]

def _parse_nac_regimen(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s or s == "なし":
        return None
    for pat, code in NAC_REGIMEN_PATTERNS:
        if re.search(pat, s, re.IGNORECASE):
            return code
    return 99


# ============================================================
# 化学療法効果 (NAC)
# ============================================================
CHEMO_EFFECT_MAP = {
    "Grade 0": 0, "Grade0": 0,
    "Grade 1a": 1, "Grade1a": 1,
    "Grade 1b": 2, "Grade1b": 2,
    "Grade 2": 3, "Grade2": 3,
    "Grade 3": 4, "Grade3": 4,
}

# ============================================================
# Palliative chemo regimen
# ============================================================
PAL_REGIMEN_PATTERNS = [
    (r"SOX", 1), (r"CAPOX", 2), (r"SP$|S-?1\+CDDP|SP療法", 3),
    (r"Nivo.*SOX", 4), (r"Nivo.*CAPOX", 5),
    (r"T-Mab|Tmab|トラスツズマブ", 6),
    (r"nab-PTX|nab.?パクリ", 8),
    (r"RAM.*PTX|ラムシルマブ", 10),
    (r"Nivolumab|ニボルマブ", 11),
    (r"Pembrolizumab|ペムブロリズマブ", 12),
    (r"TAS.?102|ロンサーフ", 13),
    (r"Irinotecan|イリノテカン|CPT", 14),
    (r"T-?DXd|エンハーツ", 15),
    (r"S-?1$|S1$|TS-?1|UFT|5-?FU|テガフール", 99),  # 古いレジメン→その他
    (r"MMC|マイトマイシン", 99),
    (r"MTX|メソトレキサート", 99),
    (r"CDDP|シスプラチン", 99),
    (r"Docetaxel|ドセタキセル|DTX", 99),
    (r"Paclitaxel|パクリタキセル|PTX", 99),
]

def _parse_pal_regimen(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s or s == "なし":
        return None
    for pat, code in PAL_REGIMEN_PATTERNS:
        if re.search(pat, s, re.IGNORECASE):
            return code
    return 99


# ============================================================
# メイン変換ロジック
# ============================================================
def convert_row(ws, row):
    c = lambda col: _cell(ws, row, col)
    record = OrderedDict()
    warnings = []

    # ------ patients ------
    record["patient_id"] = c(2)  # カルテ番号
    record["sex"] = _map_val(c(5), SEX_MAP)
    record["birthdate"] = _fmt_date(c(6))
    record["admission_date"] = _fmt_date(c(7))
    record["surgery_date"] = _fmt_date(c(8))
    record["discharge_date"] = _fmt_date(c(9))
    record["disease_category"] = 1  # 胃癌

    # 疾患分類
    dc = _map_val(c(20), DISEASE_CLASS_MAP)
    record["disease_class"] = dc if dc else 1

    # 規約バージョン: 手術年から自動判定
    #   〜2009年  → 第13版 (classification_versions.id = 4)
    #   2010〜2016年 → 第14版 (classification_versions.id = 5)
    #   2017年〜     → 第15版 (classification_versions.id = 1)
    surg_date_str = _fmt_date(c(8))
    if surg_date_str:
        try:
            surg_year = int(surg_date_str[:4])
            if surg_year <= 2009:
                record["classification_version_id"] = 4
            elif surg_year <= 2016:
                record["classification_version_id"] = 5
            else:
                record["classification_version_id"] = 1
        except (ValueError, IndexError):
            pass

    # 症状
    sym_flags = _decompose_symptoms(c(22))
    record.update(sym_flags)

    # 併存疾患
    comor_flags = _decompose_comorbidities(c(50), c(51))
    record.update(comor_flags)

    # 重複癌
    dup = c(21)
    if dup and "なし" not in str(dup):
        record["synchronous_cancer_yn"] = 1
    else:
        record["synchronous_cancer_yn"] = 0

    # HP eradication: not in old DB
    # ASA, PS: not in old DB

    # ------ tumor_preop ------
    tp = "tumor_preop"
    record[f"{tp}.c_tumor_number"] = _safe_int(c(23))
    record[f"{tp}.c_tumor_size_major_mm"] = _safe_int(c(24))
    record[f"{tp}.c_tumor_size_minor_mm"] = _safe_int(c(25))
    record[f"{tp}.c_location_long"] = c(26)  # テキスト
    record[f"{tp}.c_esophageal_invasion_mm"] = _safe_int(c(52))
    record[f"{tp}.c_egj_distance_mm"] = _safe_int(c(56))

    # 肉眼型
    record[f"{tp}.c_macroscopic_type"] = _map_val(c(35), MACRO_TYPE_MAP)
    record[f"{tp}.c_type0_subclass"] = _map_val(c(36), TYPE0_MAP)

    # 組織型
    record[f"{tp}.c_histology1"] = _map_val(c(58), HISTOLOGY_MAP)

    # cT (14版 col40 → 旧cT col38 → テキスト col37)
    record[f"{tp}.c_depth"] = _parse_ct(c(40), c(38), c(37))

    # cN (14版 col42 → 旧 col41)
    record[f"{tp}.c_ln_metastasis"] = _parse_cn(c(42), c(41))

    # cH, cP, cM
    record[f"{tp}.c_liver_metastasis"] = _map_val(c(43), CH_MAP)
    record[f"{tp}.c_peritoneal"] = _map_val(c(44), CP_MAP)
    record[f"{tp}.c_distant_metastasis"] = _map_val(c(45), CM_MAP)

    # cStage (15版 col49 → 14版 col48 → 旧 col47)
    record[f"{tp}.c_stage"] = _parse_stage(c(49), c(48), c(47), C_STAGE_MAP)

    # Preop labs (CEA, AFP, CA19-9)
    record[f"{tp}.preop_cea"] = _safe_float(c(114))
    record[f"{tp}.preop_ca199"] = _safe_float(c(116))

    # ------ neoadjuvant ------
    neo = "neoadjuvant"
    nac_yn = c(147)
    if nac_yn and str(nac_yn).strip() == "あり":
        record[f"{neo}.nac_yn"] = 1
        record[f"{neo}.nac_regimen"] = _parse_nac_regimen(c(148))
        if record[f"{neo}.nac_regimen"] == 99 and c(148):
            record[f"{neo}.nac_regimen_other"] = c(148)
        record[f"{neo}.nac_start_date"] = _fmt_date(c(149))
        # NAC effect
        eff = c(150)
        if eff:
            record[f"pathology.p_chemo_effect"] = _map_val(eff, CHEMO_EFFECT_MAP)
    else:
        record[f"{neo}.nac_yn"] = 0

    # ------ surgery ------
    surg = "surgery"
    record[f"{surg}.op_approach"] = _map_val(c(121), APPROACH_MAP)

    # 術式
    proc = _map_val(c(122), PROCEDURE_MAP)
    record[f"{surg}.op_procedure"] = proc
    if proc is None and c(122):
        record[f"{surg}.op_procedure"] = 9
        record[f"{surg}.op_procedure_other"] = c(122)

    # 郭清 (新 col126, 旧 col125)
    record[f"{surg}.op_dissection"] = _parse_dissection(c(126), c(125))

    # 再建
    record[f"{surg}.op_reconstruction"] = _map_val(c(128), RECON_MAP)
    if record[f"{surg}.op_reconstruction"] is None and c(128):
        record[f"{surg}.op_reconstruction"] = 9
        record[f"{surg}.op_reconstruction_other"] = c(128)

    # 器械吻合 (col 129) → テキストのみ参考（コードなし）

    # 合併切除
    comb_flags = _decompose_combined(c(123), c(124))
    for k, v in comb_flags.items():
        record[f"{surg}.{k}"] = v

    # 開腹移行
    conv = c(138)
    if conv and "あり" in str(conv):
        record[f"{surg}.op_conversion_yn"] = 1
    else:
        record[f"{surg}.op_conversion_yn"] = 0

    # 手術時間・出血量
    record[f"{surg}.op_time_min"] = _safe_int(c(135))
    record[f"{surg}.op_blood_loss_ml"] = _safe_int(c(137))

    # 輸血
    transfusion = c(136)
    if transfusion and str(transfusion).strip() not in ("なし", "0"):
        record[f"{surg}.op_transfusion_intra"] = 1
    else:
        record[f"{surg}.op_transfusion_intra"] = 0

    # 再手術
    reop = c(133)
    if reop and str(reop).strip() not in ("なし", "0"):
        record[f"{surg}.op_reop_yn"] = 1
    else:
        record[f"{surg}.op_reop_yn"] = 0

    # 在院日数
    record[f"{surg}.op_icu_days"] = None  # 旧DBなし

    # 合併症
    comp_flags = _decompose_complications(c(130), c(131))
    for k, v in comp_flags.items():
        record[f"{surg}.{k}"] = v

    # CD grade 推定
    record[f"{surg}.op_cd_grade_max"] = _estimate_cd_grade(c(130), c(131))

    # ------ pathology ------
    path = "pathology"
    # pT (14版 col70 → 旧pT col67 → テキスト col68)
    record[f"{path}.p_depth"] = _parse_pt(c(70), c(67), c(68))

    record[f"{path}.p_inf"] = _map_val(c(60), INF_MAP)
    # INF 14th (col61) — use if col60 empty
    if record[f"{path}.p_inf"] is None:
        record[f"{path}.p_inf"] = _map_val(c(61), INF_MAP)

    record[f"{path}.p_ly"] = _map_val(c(62), LY_MAP)
    record[f"{path}.p_v"] = _map_val(c(63), V_MAP)

    # 断端
    record[f"{path}.p_pm"] = _parse_margin(c(72))
    if record[f"{path}.p_pm"] is None:
        record[f"{path}.p_pm"] = _parse_margin(c(73))  # 14th fallback
    record[f"{path}.p_pm_mm"] = _safe_float(c(54))

    record[f"{path}.p_dm"] = _parse_margin(c(74))
    if record[f"{path}.p_dm"] is None:
        record[f"{path}.p_dm"] = _parse_margin(c(75))
    record[f"{path}.p_dm_mm"] = _safe_float(c(55))

    # pN: col143（全転移陽性個数）→ 14/15版コード。9999=NX。
    # col143 欠損時は col78（14版pN テキスト）→ col77（13版pN）にfallback。
    pn_val = _parse_pn_from_count(c(143), c(78))
    if pn_val is None:
        pn_val = _parse_pn(c(78), c(77))
    record[f"{path}.p_ln_metastasis"] = pn_val
    # 転移陽性リンパ節総数もそのまま記録
    ln_count_raw = c(143)
    if ln_count_raw is not None:
        try:
            ln_n = int(float(ln_count_raw))
            if ln_n != 9999:
                record[f"{path}.p_ln_positive_total"] = ln_n
        except (ValueError, TypeError):
            pass

    # pH, pP, pM
    record[f"{path}.p_liver"] = _map_val(c(79), PH_MAP)
    record[f"{path}.p_peritoneal"] = _map_val(c(81), PP_MAP)
    record[f"{path}.p_distant_metastasis"] = _map_val(c(83), PM_MAP)

    # CY
    record[f"{path}.p_cytology"] = _map_val(c(85), CY_MAP)

    # 遺残
    record[f"{path}.p_residual_tumor"] = _map_val(c(87), RESIDUAL_MAP)

    # pStage (15版 col89 → 14版 col88)
    record[f"{path}.p_stage"] = _parse_stage(c(89), c(88), None, P_STAGE_MAP)

    # 組織型（病理）— use same histology if separate pathology histology column exists
    # (old DB uses same col 58 for both clinical and pathological histology)
    record[f"{path}.p_histology1"] = _map_val(c(58), HISTOLOGY_MAP)

    # ------ adjuvant_chemo ------
    adj = "adjuvant_chemo"
    adj_yn = c(151)
    if adj_yn and str(adj_yn).strip() in ("あり", "UFT"):
        record[f"{adj}.adj_yn"] = 1
        record[f"{adj}.adj_start_date"] = _fmt_date(c(153))
        adj_reg_text = c(152) or str(adj_yn) if adj_yn == "UFT" else c(152)
        if adj_reg_text:
            record[f"{adj}.adj_regimen"] = 99  # 旧レジメン→その他
            record[f"{adj}.adj_regimen_other"] = str(adj_reg_text)
    else:
        record[f"{adj}.adj_yn"] = 0

    # ------ palliative_chemo ------
    pal_yn = c(154)
    if pal_yn and str(pal_yn).strip() in ("あり", "ありSOX"):
        # 1st line (col 155, 156)
        for line_n, (reg_col, date_col) in enumerate([
            (155, 156), (157, 158), (159, 160), (161, 162)
        ], start=1):
            reg_text = c(reg_col)
            if reg_text is None:
                continue
            prefix = f"palliative_chemo.line{line_n}"
            record[f"{prefix}_regimen"] = _parse_pal_regimen(reg_text)
            if record[f"{prefix}_regimen"] == 99:
                record[f"{prefix}_regimen_other"] = str(reg_text)
            record[f"{prefix}_start_date"] = _fmt_date(c(date_col))

    # ------ outcome ------
    out = "outcome"
    rec_yn = _map_val(c(13), RECURRENCE_MAP_YN)
    record[f"{out}.recurrence_yn"] = rec_yn
    record[f"{out}.recurrence_date"] = _fmt_date(c(10))
    record[f"{out}.last_alive_date"] = _fmt_date(c(11))
    record[f"{out}.death_date"] = _fmt_date(c(12))

    # 生死・死因 分離
    vital, death_cause = _parse_vital_death(c(17))
    record[f"{out}.vital_status"] = vital
    record[f"{out}.death_cause"] = death_cause

    # 再発形式分解
    rec_flags = _decompose_recurrence(c(14), c(15), c(16))
    for k, v in rec_flags.items():
        record[f"{out}.{k}"] = v

    # ------ None除去 ------
    record = OrderedDict((k, v) for k, v in record.items() if v is not None)

    return record, warnings


# ============================================================
# CSV出力
# ============================================================
def convert_excel_to_csv(excel_path, output_path=None, dry_run=False):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    max_row = ws.max_row

    print(f"📖 読み込み: {excel_path}")
    print(f"   シート: {ws.title}, {max_row - 1} 行")

    all_records = []
    all_warnings = []
    skipped = 0

    for row in range(2, max_row + 1):
        if _cell(ws, row, 2) is None and _cell(ws, row, 3) is None:
            skipped += 1
            continue

        record, warnings = convert_row(ws, row)

        if "surgery_date" not in record:
            all_warnings.append(f"行{row}: surgery_date が空 → スキップ")
            skipped += 1
            continue

        all_records.append(record)
        for w in warnings:
            all_warnings.append(f"行{row}: {w}")

    print(f"   変換成功: {len(all_records)} 行, スキップ: {skipped} 行")

    # 統合カラム
    all_cols = OrderedDict()
    for rec in all_records:
        for k in rec:
            all_cols[k] = True
    col_list = list(all_cols.keys())

    if dry_run:
        print(f"\n🔍 ドライラン: {len(col_list)} カラム")
        for c in col_list[:30]:
            print(f"   {c}")
        if len(col_list) > 30:
            print(f"   ... 他 {len(col_list) - 30} カラム")

        # 統計
        print(f"\n📊 データ統計:")
        from collections import Counter
        sex_cnt = Counter(r.get("sex") for r in all_records)
        print(f"   性別: {dict(sex_cnt)}")
        proc_cnt = Counter(r.get("surgery.op_procedure") for r in all_records)
        print(f"   術式: {dict(proc_cnt)}")
        approach_cnt = Counter(r.get("surgery.op_approach") for r in all_records)
        print(f"   アプローチ: {dict(approach_cnt)}")

        if all_warnings:
            print(f"\n⚠️  警告: {len(all_warnings)} 件")
            for w in all_warnings[:20]:
                print(f"   {w}")
        return

    if output_path is None:
        output_path = os.path.splitext(excel_path)[0] + "_ugidb.csv"

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=col_list)
        writer.writeheader()
        for rec in all_records:
            writer.writerow(rec)

    print(f"✅ CSV出力: {output_path}")
    print(f"   {len(all_records)} 行 × {len(col_list)} カラム")

    if all_warnings:
        print(f"\n⚠️  警告: {len(all_warnings)} 件")
        for w in all_warnings[:30]:
            print(f"   {w}")
        if len(all_warnings) > 30:
            print(f"   ... 他 {len(all_warnings) - 30} 件")

    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("使い方: python excel_old_to_ugidb_csv.py input.xlsx [output.csv] [--dry-run]")
        sys.exit(1)

    excel_path = sys.argv[1]
    output_path = None
    dry_run = "--dry-run" in sys.argv

    for arg in sys.argv[2:]:
        if arg != "--dry-run" and arg.endswith(".csv"):
            output_path = arg

    convert_excel_to_csv(excel_path, output_path, dry_run)
