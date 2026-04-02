"""
胃癌データベース Excel → UGI_DB CSV 変換スクリプト
=================================================
胃癌データベース_サンプル.xlsx (229列フラット構造) を
csv_import.py 互換のCSV形式に変換する。

使い方:
    python excel_to_ugidb_csv.py input.xlsx output.csv [--dry-run]
"""

import sys
import os
import re
import csv
import io
from datetime import datetime, date
from collections import OrderedDict

try:
    import openpyxl
except ImportError:
    print("openpyxl が必要です: pip install openpyxl")
    sys.exit(1)


# ============================================================
# 日付ヘルパー
# ============================================================
def _fmt_date(val):
    """datetime / 文字列 → YYYY-MM-DD。"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    # YYYY/MM/DD or YYYY-MM-DD
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y年%m月%d日"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None  # パース不能 → 警告付きスキップ


# ============================================================
# テキスト → コード変換マップ
# ============================================================

# --- 性別 ---
SEX_MAP = {"男": 1, "女": 2, "男性": 1, "女性": 2}

# --- 喫煙 ---
SMOKING_MAP = {
    "なし": 0, "禁煙": 1, "過去あり": 1, "現喫煙": 2, "現在喫煙中": 2,
    "現在喫煙": 2,
}

# --- 飲酒 ---
ALCOHOL_MAP = {
    "なし": 0, "機会飲酒": 1, "常用飲酒": 2, "常習飲酒": 2,
}

# --- ASA ---
ASA_MAP = {
    "classI": 1, "classII": 2, "classIII": 3, "classIV": 4,
    "classV": 5, "classVI": 6,
    "class1": 1, "class2": 2, "class3": 3,
    "I": 1, "II": 2, "III": 3, "IV": 4, "V": 5,
    "1": 1, "2": 2, "3": 3, "4": 4, "5": 5,
}

# --- 除菌療法 ---
HP_MAP = {
    "なし": 0, "あり": 1, "除菌成功": 1, "除菌不成功": 2,
    "未施行": 3, "不明": 3,
}

# --- 疾患分類 ---
DISEASE_CLASS_MAP = {
    "初発胃癌": 1, "初発胃癌（ESD後）": 2, "残胃の癌": 3,
    "GIST": 4, "gist": 4,
    "悪性リンパ腫（B cell）": 5, "悪性リンパ腫（T cell）": 6,
    "悪性リンパ腫（その他）": 7,
    "その他（神経原性腫瘍）": 8, "その他（平滑筋腫瘍）": 9,
    "食道癌": 10,
}

# --- 肉眼型 ---
MACRO_TYPE_MAP = {
    "0型": 0, "1型": 1, "2型": 2, "3型": 3, "4型": 4, "5型": 5,
    "Type 0": 0, "Type 1": 1, "Type 2": 2, "Type 3": 3,
    "Type 4": 4, "Type 5": 5,
}

# --- 0型亜分類 ---
TYPE0_MAP = {
    "I": 1, "IIa": 2, "IIb": 3, "IIc": 4, "III": 5,
    "0-I": 1, "0-IIa": 2, "0-IIb": 3, "0-IIc": 4, "0-III": 5,
    "Ip": 1, "Is": 1, "Isp": 1,
}

# --- 組織型 (胃癌) ---
HISTOLOGY_MAP = {
    "tub1": 1, "tub2": 2, "por1": 3, "por2": 4, "por": 3,
    "sig": 5, "muc": 6, "pap": 7,
    "neuroendocrine carcinoma": 8, "adenosquamous": 9,
    "NEC": 8,
}

# --- cT (胃癌規約15版) ---
def _parse_ct_gastric(val):
    if val is None:
        return None
    s = str(val).strip()
    patterns = [
        (r"cT1a|cT1a\s*\(M\)|M$", 1),
        (r"cT1b|cT1b\d?\s*\(SM\d?\)|SM", 2),
        (r"cT2\s*\(MP\)|cT2|MP$", 3),
        (r"cT3\s*\(SS\)|cT3|SS$", 4),
        (r"cT4a\s*\(SE\)|cT4a|SE$", 5),
        (r"cT4b\s*\(SI\)|cT4b|SI$", 6),
        (r"cTX|TX", 9),
    ]
    for pat, code in patterns:
        if re.search(pat, s, re.IGNORECASE):
            return code
    return None

# --- pT (胃癌規約15版) ---
def _parse_pt_gastric(val):
    if val is None:
        return None
    s = str(val).strip()
    patterns = [
        (r"pT0$", 0),
        (r"pT1a|pT1a\s*\(M\)", 1),
        (r"pT1b1\s*\(SM1\)|pT1b\s*\(SM\)|pT1b1|pT1b2|pT1b", 2),
        (r"pT2\s*\(MP\)|pT2", 3),
        (r"pT3\s*\(SS\)|pT3", 4),
        (r"pT4a\s*\(SE\)|pT4a", 5),
        (r"pT4b\s*\(SI\)|pT4b", 6),
        (r"pTX|pTx", 9),
    ]
    for pat, code in patterns:
        if re.search(pat, s, re.IGNORECASE):
            return code
    return None

# --- cN (胃癌) ---
CN_MAP = {
    "cN0": 0, "cN+": 1, "cN1": 1, "cN2": 2, "cN3": 3,
    "cN3a": 3, "cN3b": 4, "cNX": 9,
}

# --- pN (胃癌) ---
PN_MAP = {
    "pN0": 0, "pN1": 1, "pN2": 2, "pN3": 3,
    "pN3a": 3, "pN3b": 4, "pNX": 9,
}

# --- cM ---
CM_MAP = {"cM0": 0, "cM1": 1, "cMX": 9}

# --- pM ---
PM_META_MAP = {"pM0": 0, "pMO": 0, "pM1": 1, "pMX": 9}  # pMO=typo in Excel

# --- cP (腹膜) ---
CP_MAP = {"cP0": 0, "cP1": 1}

# --- pP (腹膜) ---
PP_MAP = {"pP0": 0, "pP1": 1}

# --- cH / pH (肝転移) ---
CH_MAP = {"cH0": 0, "cH1": 1}
PH_MAP = {"pH0": 0, "pH1": 1}

# --- CY (洗浄細胞診) ---
CY_MAP = {
    "pCY0": 0, "CY0": 0,
    "pCY1": 1, "CY1": 1,
    "pCYX": 9, "CYX": 9,
}
def _parse_cy(val):
    if val is None:
        return None
    s = str(val).strip()
    for key, code in CY_MAP.items():
        if s.startswith(key):
            return code
    if "施行せず" in s or "未施行" in s:
        return 9
    return None

# --- cStage (胃癌) ---
C_STAGE_MAP = {
    "cI": 1, "cIA": 1, "cIB": 1,
    "cII": 2, "cIIA": 2, "cIIB": 3,
    "cIII": 4, "cIIIA": 4, "cIIIB": 4, "cIIIC": 4,
    "cIV": 5, "cIVA": 5, "cIVB": 6,
}

# --- pStage (胃癌) ---
P_STAGE_MAP = {
    "p0": 0, "pIA": 1, "pIB": 2,
    "pIIA": 3, "pIIB": 4,
    "pIIIA": 5, "pIIIB": 6, "pIIIC": 7,
    "pIV": 8,
    # short forms
    "pI": 1, "pII": 3, "pIII": 5,
}

# --- 到達法 (op_approach) ---
APPROACH_MAP = {
    "開腹": 1, "腹腔鏡下": 2, "腹腔鏡": 2,
    "ロボット支援下": 3, "ロボット支援": 3, "ロボット": 3,
    "開胸": 4, "胸腔鏡": 5, "HALS": 6, "腹腔鏡補助": 7,
}

# --- 術式 (op_procedure_gastric) ---
PROCEDURE_MAP = {
    "幽門側胃切除": 1, "DG": 1,
    "胃全摘": 2, "TG": 2,
    "噴門側胃切除": 3, "PG": 3,
    "胃局所切除": 4, "局所切除": 4,
    "幽門保存胃切除": 5, "PPG": 5,
    "残胃全摘": 6,
    "審査腹腔鏡": 7, "staging laparoscopy": 7,
    "バイパス手術": 8, "バイパス": 8,
}

# --- 郭清 (op_dissection_gastric) ---
DISSECTION_MAP = {
    "D0": 0, "D1": 1, "D1+": 2, "D2": 3, "D2+": 4, "D3": 5,
}

# --- 再建 (op_reconstruction_gastric) ---
RECON_MAP = {
    "B-I": 1, "B-1": 1, "BI": 1, "Billroth I": 1,
    "B-II": 2, "B-2": 2, "BII": 2, "Billroth II": 2,
    "Roux-en-Y": 3, "R-Y": 3, "RY": 3, "Roux en Y": 3,
    "食道残胃吻合": 4,
    "ダブルトラクト": 5, "double tract": 5,
    "空腸間置": 6, "JI": 6,
    "なし": 7,
}

# --- 吻合法 (op_anastomosis_method) ---
ANAST_MAP = {
    "三角吻合": 1, "デルタ吻合": 2,
    "T吻合": 3, "Overlap+Side-to-side": 3,
    "DST": 4, "FEEA": 5, "Overlap": 6,
    "上川法": 7, "上川法（観音開き法）": 7, "観音開き法": 7,
    "mSOFY法": 8, "mSOFY": 8,
    "OrVil": 9, "手縫い吻合": 10,
    "端端吻合": 11, "端側吻合": 12,
}

# --- 蠕動方向 ---
PERISTALSIS_MAP = {"順蠕動": 1, "逆蠕動": 2}

# --- 再建経路 ---
RECON_ROUTE_MAP = {"後縦隔": 1, "胸骨後": 2, "胸壁前": 3}

# --- 完遂 ---
COMPLETION_MAP = {
    "予定通り完遂": 1, "完遂": 1,
    "コンバージョン": 2, "conversion": 2,
    "中止": 3,
}

# --- CD分類 ---
CD_GRADE_MAP = {
    "Grade0": 0, "Grade 0": 0, "なし": 0,
    "GradeI": 1, "Grade I": 1, "Grade1": 1,
    "GradeII": 2, "Grade II": 2, "Grade2": 2,
    "GradeIIIa": 3, "Grade IIIa": 3, "Grade3a": 3,
    "GradeIIIb": 4, "Grade IIIb": 4, "Grade3b": 4,
    "GradeIVa": 5, "Grade IVa": 5, "Grade4a": 5,
    "GradeIVb": 6, "Grade IVb": 6, "Grade4b": 6,
    "GradeV": 7, "Grade V": 7, "死亡": 7,
}

# --- INF (浸潤増殖様式) ---
INF_MAP = {"INFa": 1, "INFb": 2, "INFc": 3}

# --- ly (リンパ管侵襲) ---
LY_MAP = {"Ly0": 0, "ly0": 0, "Ly1a": 1, "ly1a": 1,
           "Ly1b": 2, "ly1b": 2, "Ly2": 3, "ly2": 3, "Ly3": 3}

# --- v (静脈侵襲) ---
V_MAP = {"V0": 0, "v0": 0, "V1a": 1, "v1a": 1,
          "V1b": 2, "v1b": 2, "V2": 3, "v2": 3, "V3": 3}

# --- PM (口側断端) ---
PM_MARGIN_MAP = {"pPM0": 0, "PM0": 0, "pPM1": 1, "PM1": 1}

# --- DM (肛側断端) ---
DM_MAP = {"pDM0": 0, "DM0": 0, "pDM1": 1, "DM1": 1}

# --- 遺残 ---
RESIDUAL_MAP = {
    "pR0": 0, "R0": 0,
    "pR1": 1, "R1": 1,
    "pR2": 2, "R2": 2,
    "pRX": 8, "RX": 8,
}
def _parse_residual(val):
    if val is None:
        return None
    s = str(val).strip()
    for key, code in RESIDUAL_MAP.items():
        if s.startswith(key):
            return code
    if "遺残がない" in s:
        return 0
    if "顕微鏡的遺残" in s:
        return 1
    if "肉眼的遺残" in s:
        return 2
    return None

# --- 再発有無 ---
RECURRENCE_MAP = {"再発": 1, "あり": 1, "無再発": 0, "なし": 0, "不明": 9}

# --- 生死 → vital_status ---
VITAL_MAP = {
    "生存中": 1, "生存": 1,
    "死亡": 2, "原病死": 2,
    "他病死": 3, "手術関連死": 4, "事故死": 5,
    "不明": 9,
}

# --- 死因 ---
DEATH_CAUSE_MAP = {
    "原病死": 1, "原発巣増悪": 1,
    "腹膜播種": 2, "肝転移": 3, "肺転移": 4,
    "手術関連死": 6, "他病死": 7,
    "事故死": 11,
}

# --- MSI ---
MSI_MAP = {"MSS": 0, "MSI-H": 1, "MSI-High": 1, "MSI-L": 2, "MSI-Low": 2, "未検": 9}

# --- HER2 ---
HER2_MAP = {
    "陰性": 0, "0": 0, "1+": 0,
    "2+": 1, "Equivocal": 1,
    "陽性": 2, "3+": 2,
    "未検": 9,
}

# --- PD-L1 ---
PDL1_MAP = {"陰性": 0, "陽性": 1, "未検": 9}

# --- GIST免疫染色 ---
GIST_IHC_MAP = {"陰性": 0, "陽性": 1, "未検": 9}

# --- Fletcher ---
FLETCHER_MAP = {
    "Very low risk": 1, "very low": 1,
    "Low risk": 2, "low": 2,
    "Intermediate risk": 3, "intermediate": 3,
    "High risk": 4, "high": 4,
}

# --- 術前療法: あり/なし ---
YN_MAP = {"なし": 0, "あり": 1}

# --- 術後化学療法レジメン (adj) ---
ADJ_REGIMEN_MAP = {
    "S-1": 1, "S1": 1,
    "CAPOX": 2, "SOX": 3, "DS": 4,
    "Nivo+CAPOX": 5, "Nivo+SOX": 6,
}

# --- 化学療法完遂 ---
CHEMO_COMP_MAP = {"完遂": 1, "減量完遂": 2, "中止": 3}

# --- 薬物放射線治療効果 ---
CHEMO_EFFECT_MAP = {
    "Grade 0": 0, "Grade0": 0,
    "Grade 1a": 1, "Grade1a": 1,
    "Grade 1b": 2, "Grade1b": 2,
    "Grade 2": 3, "Grade2": 3,
    "Grade 3": 4, "Grade3": 4,
}

# --- RECIST ---
RECIST_MAP = {"CR": 1, "PR": 2, "SD": 3, "PD": 4, "NE": 5}

# --- Palliative chemo regimen (fuzzy match) ---
PAL_REGIMEN_PATTERNS = [
    (r"SOX", 1), (r"CAPOX", 2), (r"SP$", 3),
    (r"Nivo.*SOX|SOX.*Nivo", 4), (r"Nivo.*CAPOX|CAPOX.*Nivo", 5),
    (r"T-Mab.*CAPOX|トラスツズマブ.*CAPOX", 6),
    (r"T-Mab.*SOX|トラスツズマブ.*SOX", 7),
    (r"nab-PTX|nab.?パクリ", 8),
    (r"RAM.*nab|ラムシルマブ.*nab", 9),
    (r"RAM.*PTX|ラムシルマブ.*PTX", 10),
    (r"Nivolumab|ニボルマブ", 11),
    (r"Pembrolizumab|ペムブロリズマブ", 12),
    (r"TAS.?102|ロンサーフ", 13),
    (r"Irinotecan|イリノテカン|CPT-11", 14),
    (r"T-DXd|エンハーツ", 15),
    (r"Zolbetuximab|ゾルベツキシマブ", 16),
]
def _parse_pal_regimen(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s or s == "なし":
        return None
    # Try patterns in order (more specific first handled by order above)
    for pat, code in PAL_REGIMEN_PATTERNS:
        if re.search(pat, s, re.IGNORECASE):
            return code
    return 99  # その他


# ============================================================
# 併存疾患テキスト → フラグ分解
# ============================================================
COMORBIDITY_KEYWORDS = {
    "comor_hypertension": ["高血圧", "HT", "hypertension"],
    "comor_cardiovascular": ["心疾患", "心不全", "不整脈", "弁膜症", "虚血性心疾患",
                             "狭心症", "心筋梗塞", "大動脈瘤", "PAD"],
    "comor_cerebrovascular": ["脳梗塞", "脳出血", "脳卒中", "TIA", "くも膜下"],
    "comor_respiratory": ["喘息", "COPD", "間質性肺炎", "肺気腫", "呼吸器"],
    "comor_renal": ["腎不全", "CKD", "腎機能障害", "腎疾患"],
    "comor_renal_dialysis": ["透析", "HD", "dialysis"],
    "comor_hepatic": ["肝硬変", "肝炎", "肝障害", "B型肝炎", "C型肝炎"],
    "comor_diabetes": ["糖尿病", "DM", "diabetes"],
    "comor_endocrine": ["甲状腺", "副腎", "内分泌"],
    "comor_collagen": ["膠原病", "リウマチ", "SLE", "強皮症"],
    "comor_hematologic": ["貧血", "白血病", "血液疾患", "骨髄"],
    "comor_neurologic": ["パーキンソン", "てんかん", "認知症", "神経疾患"],
    "comor_psychiatric": ["うつ", "統合失調", "精神", "不安障害"],
}


def _decompose_comorbidities(vals):
    """併存疾患1〜4 → comor_* フラグ辞書。"""
    flags = {k: 0 for k in COMORBIDITY_KEYWORDS}
    combined = " ".join(str(v) for v in vals if v and str(v).strip() != "該当なし")
    if not combined.strip():
        return flags
    for col, keywords in COMORBIDITY_KEYWORDS.items():
        for kw in keywords:
            if kw in combined:
                flags[col] = 1
                break
    return flags


# ============================================================
# 内服薬テキスト → フラグ分解
# ============================================================
MEDICATION_KEYWORDS = {
    "med_antihypertensive": ["降圧", "ARB", "ACE", "Ca拮抗", "利尿"],
    "med_antithrombotic": ["抗血栓", "ワーファリン", "バイアスピリン", "抗凝固",
                           "クロピドグレル", "DOAC", "抗血小板"],
    "med_oral_hypoglycemic": ["経口血糖降下", "メトホルミン", "DPP-4"],
    "med_insulin": ["インスリン", "insulin"],
    "med_steroid_immunosup": ["ステロイド", "免疫抑制", "プレドニン", "タクロリムス"],
    "med_antineoplastic": ["抗がん", "化学療法"],
    "med_thyroid": ["チラーヂン", "甲状腺ホルモン"],
    "med_psychotropic": ["向精神", "抗うつ", "睡眠", "抗不安"],
}


def _decompose_medications(vals):
    """内服薬1〜4 → med_* フラグ辞書。"""
    flags = {k: 0 for k in MEDICATION_KEYWORDS}
    combined = " ".join(str(v) for v in vals if v and str(v).strip() not in ("該当なし", "なし"))
    if not combined.strip():
        return flags
    for col, keywords in MEDICATION_KEYWORDS.items():
        for kw in keywords:
            if kw in combined:
                flags[col] = 1
                break
    return flags


# ============================================================
# 症状テキスト → フラグ分解
# ============================================================
SYMPTOM_KEYWORDS = {
    "sym_asymptomatic": ["なし", "無症状", "検診発見", "検診"],
    "sym_epigastric_pain": ["腹痛", "心窩部痛", "上腹部痛", "胃痛"],
    "sym_dysphagia": ["嚥下障害", "嚥下困難", "つかえ"],
    "sym_weight_loss": ["体重減少", "るいそう"],
    "sym_anemia": ["貧血"],
    "sym_melena": ["黒色便", "下血", "メレナ"],
    "sym_hematemesis": ["吐血", "嘔吐", "黒色嘔吐"],
    "sym_nausea_vomiting": ["嘔気", "悪心", "嘔吐"],
    "sym_abdominal_distension": ["腹部膨満", "腹満"],
    "sym_obstruction": ["閉塞", "通過障害", "狭窄"],
}


def _decompose_symptoms(val):
    """症状テキスト → sym_* フラグ辞書。"""
    flags = {k: 0 for k in SYMPTOM_KEYWORDS}
    if val is None:
        return flags
    s = str(val).strip()
    if not s:
        return flags
    # 「なし」単独の場合は asymptomatic
    if s in ("なし", "なし（検診発見）", "無症状"):
        flags["sym_asymptomatic"] = 1
        return flags
    for col, keywords in SYMPTOM_KEYWORDS.items():
        for kw in keywords:
            if kw in s:
                flags[col] = 1
                break
    return flags


# ============================================================
# 合併症テキスト → comp_* フラグ分解
# ============================================================
COMPLICATION_KEYWORDS = {
    "comp_ssi": ["SSI", "創感染", "創部感染"],
    "comp_wound_dehiscence": ["創離開", "wound dehiscence"],
    "comp_intra_abd_abscess": ["腹腔内膿瘍", "膿瘍"],
    "comp_bleeding": ["出血", "bleeding"],
    "comp_ileus": ["イレウス", "腸閉塞", "ileus"],
    "comp_dvt_pe": ["DVT", "PE", "肺塞栓", "深部静脈血栓"],
    "comp_pneumonia": ["肺炎", "pneumonia"],
    "comp_atelectasis": ["無気肺", "atelectasis"],
    "comp_uti": ["尿路感染", "UTI"],
    "comp_delirium": ["せん妄", "delirium"],
    "comp_cardiac": ["不整脈", "心不全", "心房細動", "cardiac"],
    "comp_dge": ["胃排泄遅延", "DGE", "delayed gastric emptying"],
    "comp_perforation": ["穿孔", "perforation"],
    "comp_cholelithiasis": ["胆石", "胆嚢炎"],
    "comp_anastomotic_leak": ["縫合不全", "吻合部漏", "anastomotic leak"],
    "comp_anastomotic_stricture": ["吻合部狭窄", "anastomotic stricture"],
    "comp_anastomotic_bleeding": ["吻合部出血"],
    "comp_pancreatic_fistula": ["膵液瘻", "膵瘻", "POPF", "pancreatic fistula"],
    "comp_bile_leak": ["胆汁漏", "bile leak"],
    "comp_duodenal_stump_leak": ["十二指腸断端", "duodenal stump"],
    "comp_rln_palsy": ["反回神経麻痺", "嗄声", "RLN"],
    "comp_chylothorax": ["乳糜胸", "chylothorax"],
    "comp_empyema": ["膿胸", "empyema"],
    "comp_pneumothorax": ["気胸", "pneumothorax"],
    "comp_ards": ["ARDS", "急性呼吸促迫"],
    "comp_dic": ["DIC", "播種性血管内凝固"],
    "comp_sepsis": ["敗血症", "sepsis"],
    "comp_renal_failure": ["腎不全", "renal failure", "AKI"],
    "comp_hepatic_failure": ["肝不全", "hepatic failure"],
}


def _decompose_complications(comp1, comp2):
    """合併症テキスト1,2 → comp_* フラグ辞書 + op_complication_yn。"""
    flags = {k: 0 for k in COMPLICATION_KEYWORDS}
    combined = " ".join(str(v) for v in [comp1, comp2] if v and str(v).strip() != "なし")
    if not combined.strip():
        flags["op_complication_yn"] = 0
        return flags
    flags["op_complication_yn"] = 1
    for col, keywords in COMPLICATION_KEYWORDS.items():
        for kw in keywords:
            if kw.lower() in combined.lower():
                flags[col] = 1
                break
    return flags


# ============================================================
# 再発形式テキスト → rec_* フラグ分解
# ============================================================
RECURRENCE_SITE_KEYWORDS = {
    "rec_peritoneal": ["腹膜再発", "腹膜播種", "腹膜"],
    "rec_liver": ["肝再発", "肝転移"],
    "rec_lung": ["肺再発", "肺転移"],
    "rec_lymph_node": ["リンパ節再発", "リンパ節転移"],
    "rec_local": ["局所再発", "断端再発"],
    "rec_bone": ["骨再発", "骨転移"],
    "rec_brain": ["脳再発", "脳転移"],
    "rec_ovary": ["卵巣再発", "卵巣転移", "Krukenberg"],
    "rec_adrenal": ["副腎再発", "副腎転移"],
}


def _decompose_recurrence_sites(val):
    """再発形式テキスト → rec_* フラグ辞書。"""
    flags = {k: 0 for k in RECURRENCE_SITE_KEYWORDS}
    if val is None:
        return flags
    s = str(val).strip()
    if not s or s == "なし":
        return flags
    for col, keywords in RECURRENCE_SITE_KEYWORDS.items():
        for kw in keywords:
            if kw in s:
                flags[col] = 1
                break
    return flags


# ============================================================
# 合併切除臓器テキスト → comb_* フラグ分解
# ============================================================
COMBINED_RESECTION_KEYWORDS = {
    "comb_splenectomy": ["脾臓", "脾摘", "splenectomy"],
    "comb_cholecystectomy": ["胆嚢", "cholecystectomy"],
    "comb_distal_pancreatectomy": ["膵体尾部", "膵切除"],
    "comb_transverse_colectomy": ["横行結腸", "結腸切除"],
    "comb_partial_hepatectomy": ["肝部分切除", "肝切除"],
    "comb_diaphragm": ["横隔膜"],
    "comb_adrenalectomy": ["副腎"],
    "comb_ovary": ["卵巣"],
    "comb_small_intestine": ["小腸"],
}


def _decompose_combined_resection(val):
    """合併切除テキスト → comb_* フラグ辞書。"""
    flags = {k: 0 for k in COMBINED_RESECTION_KEYWORDS}
    if val is None:
        return flags
    s = str(val).strip()
    if not s or s == "なし":
        return flags
    for col, keywords in COMBINED_RESECTION_KEYWORDS.items():
        for kw in keywords:
            if kw in s:
                flags[col] = 1
                break
    return flags


# ============================================================
# 癌家族歴テキスト → fhx_* フラグ分解
# ============================================================
FHX_KEYWORDS = {
    "fhx_gastric": ["胃癌", "胃がん"],
    "fhx_esophageal": ["食道癌", "食道がん"],
    "fhx_colorectal": ["大腸癌", "大腸がん", "直腸癌", "結腸癌"],
    "fhx_lung": ["肺癌", "肺がん"],
    "fhx_liver": ["肝癌", "肝がん", "肝臓癌"],
    "fhx_pancreas": ["膵癌", "膵がん", "膵臓癌"],
    "fhx_breast": ["乳癌", "乳がん"],
}


def _decompose_family_history(val):
    """癌家族歴テキスト → fhx_* フラグ辞書。"""
    flags = {k: 0 for k in FHX_KEYWORDS}
    if val is None:
        return flags
    s = str(val).strip()
    if not s or s == "なし":
        return flags
    for col, keywords in FHX_KEYWORDS.items():
        for kw in keywords:
            if kw in s:
                flags[col] = 1
                break
    return flags


# ============================================================
# 初診時遠隔転移部位テキスト → c_meta_* フラグ分解
# ============================================================
C_META_KEYWORDS = {
    "c_meta_peritoneal": ["腹膜"],
    "c_meta_liver": ["肝"],
    "c_meta_lung": ["肺"],
    "c_meta_lymph_node": ["リンパ節"],
    "c_meta_bone": ["骨"],
    "c_meta_brain": ["脳"],
    "c_meta_ovary": ["卵巣"],
}


# ============================================================
# リンパ節マッピング  No.X(M/L) → lymph_nodes.ln_X_m/l
# ============================================================
LN_STATION_MAP = {
    "1": "1", "2": "2", "3a": "3a", "3b": "3b",
    "4sa": "4sa", "4sb": "4sb", "4d": "4d",
    "5": "5", "6": "6", "7": "7",
    "8a": "8a", "8p": "8p", "9": "9",
    "10": "10", "11p": "11p", "11d": "11d",
    "12a": "12a", "14v": "14v", "16": "16",
    "19": "19", "20": "20",
    "108": "108", "110": "110", "111": "111", "112": "112",
}


# ============================================================
# メイン変換ロジック
# ============================================================
def _cell(ws, row, col):
    """セル値を取得（None安全）。"""
    v = ws.cell(row=row, column=col).value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v in ("", "該当なし", "-", "−", "ー"):
            return None
    return v


def _map_val(val, mapping, default=None):
    """テキスト値を辞書マッピング。"""
    if val is None:
        return default
    s = str(val).strip()
    if s in mapping:
        return mapping[s]
    # 部分一致も試す
    for key, code in mapping.items():
        if key in s:
            return code
    return default


def _safe_int(val):
    """数値をintに。"""
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _safe_float(val):
    """数値をfloatに。"""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def convert_row(ws, row):
    """1行をUGI_DB CSV辞書に変換する。"""
    c = lambda col: _cell(ws, row, col)
    record = OrderedDict()
    warnings = []

    # ------ patients ------
    record["patient_id"] = c(2)  # 患者ID
    record["sex"] = _map_val(c(4), SEX_MAP)
    record["birthdate"] = _fmt_date(c(5))
    record["first_visit_date"] = _fmt_date(c(6))
    record["admission_date"] = _fmt_date(c(7))
    record["surgery_date"] = _fmt_date(c(8))
    record["discharge_date"] = _fmt_date(c(9))
    record["height_cm"] = _safe_float(c(19))
    record["weight_admission"] = _safe_float(c(20))
    record["weight_discharge"] = _safe_float(c(21))
    record["smoking"] = _map_val(c(23), SMOKING_MAP)
    record["alcohol"] = _map_val(c(24), ALCOHOL_MAP)
    record["ps"] = _safe_int(c(25))
    record["asa"] = _map_val(c(26), ASA_MAP)
    record["hp_eradication"] = _map_val(c(38), HP_MAP)
    record["disease_category"] = 1  # 胃癌

    # 疾患分類
    dc_text = c(39)
    dc = _map_val(dc_text, DISEASE_CLASS_MAP)
    if dc is None and dc_text:
        # fuzzy
        if "ESD" in str(dc_text):
            dc = 2
        elif "残胃" in str(dc_text):
            dc = 3
        elif "GIST" in str(dc_text).upper():
            dc = 4
        else:
            dc = 1
            warnings.append(f"疾患分類不明: {dc_text} → 初発胃癌(1)に仮設定")
    record["disease_class"] = dc

    # 症状分解
    sym_flags = _decompose_symptoms(c(22))
    record.update(sym_flags)

    # 併存疾患分解
    comor_flags = _decompose_comorbidities([c(27), c(28), c(29), c(30)])
    record.update(comor_flags)

    # 内服薬分解
    med_flags = _decompose_medications([c(31), c(32), c(33), c(34)])
    record.update(med_flags)

    # 同時性重複癌
    sync = c(35)
    record["synchronous_cancer_yn"] = 0 if (sync is None or sync == "なし") else 1

    # 異時性重複癌
    meta_c = c(36)
    record["metachronous_cancer_yn"] = 0 if (meta_c is None or meta_c == "なし") else 1

    # 癌家族歴
    fhx_flags = _decompose_family_history(c(37))
    record.update(fhx_flags)

    # ------ tumor_preop ------
    tp = "tumor_preop"
    # 残胃情報
    if dc == 3:  # 残胃の癌
        record[f"{tp}.remnant_stomach_yn"] = 1
    else:
        record[f"{tp}.remnant_stomach_yn"] = 0

    record[f"{tp}.c_tumor_number"] = _safe_int(c(43))
    # location (テキスト→code, 簡易マッピング)
    loc_long_map = {"U": 1, "M": 2, "L": 3, "E": 4, "D": 5}
    record[f"{tp}.c_location_long"] = c(44)  # そのままテキスト
    record[f"{tp}.c_location_short"] = c(45)

    # EGJ
    egj_map = {"Siewert I": 1, "Siewert II": 2, "Siewert III": 3,
               "SiewertI": 1, "SiewertII": 2, "SiewertIII": 3}
    record[f"{tp}.c_location_egj"] = _map_val(c(46), egj_map)
    record[f"{tp}.c_egj_distance_mm"] = _safe_int(c(47))
    record[f"{tp}.c_esophageal_invasion_mm"] = _safe_int(c(48))

    # 肉眼型
    record[f"{tp}.c_macroscopic_type"] = _map_val(c(49), MACRO_TYPE_MAP)
    record[f"{tp}.c_type0_subclass"] = _map_val(c(50), TYPE0_MAP)

    record[f"{tp}.c_tumor_size_major_mm"] = _safe_int(c(51))
    record[f"{tp}.c_tumor_size_minor_mm"] = _safe_int(c(52))

    # 組織型
    hist = _map_val(c(53), HISTOLOGY_MAP)
    record[f"{tp}.c_histology1"] = hist

    # cT
    record[f"{tp}.c_depth"] = _parse_ct_gastric(c(54))

    # 浸潤臓器 (テキスト → c_inv_* フラグ, 簡略化)
    inv_text = c(55)
    # TODO: 浸潤臓器テキストの詳細分解（将来対応）

    # cN
    record[f"{tp}.c_ln_metastasis"] = _map_val(c(56), CN_MAP)

    # cM
    record[f"{tp}.c_distant_metastasis"] = _map_val(c(57), CM_MAP)

    # 遠隔転移部位
    meta_site = c(58)
    # c_meta_* フラグは将来対応

    # cP
    record[f"{tp}.c_peritoneal"] = _map_val(c(59), CP_MAP)

    # cH
    record[f"{tp}.c_liver_metastasis"] = _map_val(c(60), CH_MAP)

    # cStage
    record[f"{tp}.c_stage"] = _map_val(c(61), C_STAGE_MAP)

    # Preop labs (columns not in sample 229 columns, skip)

    # ------ neoadjuvant ------
    neo = "neoadjuvant"
    nac_yn = c(62)
    if nac_yn and nac_yn != "なし":
        record[f"{neo}.nac_yn"] = 1
        record[f"{neo}.nac_regimen"] = _map_val(c(63), {
            "SOX": 1, "SP": 2, "DOS": 3, "FLOT": 4,
            "Nivo+SOX": 5, "Nivo+CAPOX": 6,
        }, 99)
        record[f"{neo}.nac_start_date"] = _fmt_date(c(64))
        record[f"{neo}.nac_courses"] = _safe_int(c(65))
        record[f"{neo}.nac_completion"] = _map_val(c(66), CHEMO_COMP_MAP)
        record[f"{neo}.nac_adverse_event"] = c(67)
        # RECIST
        record[f"{neo}.recist_target1"] = c(68)
        record[f"{neo}.recist_target2"] = c(69)
        record[f"{neo}.recist_target3"] = c(70)
        record[f"{neo}.recist_shrinkage_pct"] = _safe_float(c(71))
        record[f"{neo}.recist_target_response"] = _map_val(c(72), RECIST_MAP)
        record[f"{neo}.recist_nontarget1"] = c(73)
        record[f"{neo}.recist_nontarget2"] = c(74)
        record[f"{neo}.recist_nontarget3"] = c(75)
        record[f"{neo}.recist_nontarget_response"] = _map_val(c(76), RECIST_MAP)
        record[f"{neo}.recist_new_lesion"] = _map_val(c(77), YN_MAP)
        record[f"{neo}.recist_new_lesion_detail"] = c(78)
        record[f"{neo}.recist_overall"] = _map_val(c(79), RECIST_MAP)
        # 胃原発巣
        record[f"{neo}.primary_shrinkage_pct"] = _safe_float(c(80))
        record[f"{neo}.primary_elevation"] = _map_val(c(81), RECIST_MAP)
        record[f"{neo}.primary_depression"] = _map_val(c(82), RECIST_MAP)
        record[f"{neo}.primary_stenosis"] = _map_val(c(83), RECIST_MAP)
        record[f"{neo}.primary_overall_response"] = _map_val(c(84), RECIST_MAP)
    else:
        record[f"{neo}.nac_yn"] = 0

    # ------ surgery ------
    surg = "surgery"
    record[f"{surg}.op_approach"] = _map_val(c(85), APPROACH_MAP)
    record[f"{surg}.op_completion"] = _map_val(c(86), COMPLETION_MAP)
    # col 87: 胸部操作 → eso_surgery 用、胃癌ではスキップ
    record[f"{surg}.op_procedure"] = _map_val(c(88), PROCEDURE_MAP)
    if record[f"{surg}.op_procedure"] is None and c(88):
        record[f"{surg}.op_procedure_other"] = c(88)
        record[f"{surg}.op_procedure"] = 9  # その他
    record[f"{surg}.op_dissection"] = _map_val(c(89), DISSECTION_MAP)
    record[f"{surg}.op_reconstruction"] = _map_val(c(90), RECON_MAP)
    if record[f"{surg}.op_reconstruction"] is None and c(90):
        record[f"{surg}.op_reconstruction_other"] = c(90)
        record[f"{surg}.op_reconstruction"] = 9
    record[f"{surg}.op_anastomosis_method"] = _map_val(c(91), ANAST_MAP)
    if record[f"{surg}.op_anastomosis_method"] is None and c(91):
        record[f"{surg}.op_anastomosis_method_other"] = c(91)
        record[f"{surg}.op_anastomosis_method"] = 99
    record[f"{surg}.op_peristalsis_direction"] = _map_val(c(92), PERISTALSIS_MAP)
    record[f"{surg}.op_reconstruction_route"] = _map_val(c(93), RECON_ROUTE_MAP)

    # 合併切除
    comb_flags = _decompose_combined_resection(c(94))
    for k, v in comb_flags.items():
        record[f"{surg}.{k}"] = v

    record[f"{surg}.op_conversion_yn"] = _map_val(c(95), {"なし": 0, "あり": 1}, 0)
    record[f"{surg}.op_time_min"] = _safe_int(c(96))
    record[f"{surg}.op_console_time_min"] = _safe_int(c(97))
    record[f"{surg}.op_blood_loss_ml"] = _safe_int(c(98))
    record[f"{surg}.op_transfusion_intra"] = _map_val(c(99), YN_MAP)
    record[f"{surg}.op_transfusion_post"] = _map_val(c(100), YN_MAP)
    record[f"{surg}.op_icu_days"] = _safe_int(c(101))
    record[f"{surg}.op_reop_yn"] = _map_val(c(102), YN_MAP)

    # 合併症
    comp_flags = _decompose_complications(c(104), c(105))
    for k, v in comp_flags.items():
        record[f"{surg}.{k}"] = v

    # CD grade
    record[f"{surg}.op_cd_grade_max"] = _map_val(c(106), CD_GRADE_MAP)

    # ------ pathology ------
    path = "pathology"
    record[f"{path}.p_tumor_number"] = _safe_int(c(107))
    record[f"{path}.p_location_long"] = c(108)
    record[f"{path}.p_location_short"] = c(109)
    record[f"{path}.p_location_egj"] = _map_val(c(110), egj_map)
    record[f"{path}.p_egj_distance_mm"] = _safe_int(c(111))
    record[f"{path}.p_esoph_invasion_mm"] = _safe_int(c(112))
    record[f"{path}.p_macroscopic_type"] = _map_val(c(113), MACRO_TYPE_MAP)
    record[f"{path}.p_type0_subclass"] = _map_val(c(114), TYPE0_MAP)
    record[f"{path}.p_size_major_mm"] = _safe_int(c(115))
    record[f"{path}.p_size_minor_mm"] = _safe_int(c(116))
    record[f"{path}.p_histology1"] = _map_val(c(117), HISTOLOGY_MAP)
    record[f"{path}.p_histology2"] = _map_val(c(118), HISTOLOGY_MAP)
    record[f"{path}.p_histology3"] = _map_val(c(119), HISTOLOGY_MAP)
    record[f"{path}.p_depth"] = _parse_pt_gastric(c(120))

    # 浸潤臓器 (col 121): テキスト → p_inv_* フラグ (将来対応)

    record[f"{path}.p_inf"] = _map_val(c(122), INF_MAP)
    record[f"{path}.p_ly"] = _map_val(c(123), LY_MAP)
    record[f"{path}.p_v"] = _map_val(c(124), V_MAP)
    record[f"{path}.p_pm"] = _map_val(c(125), PM_MARGIN_MAP)
    record[f"{path}.p_pm_mm"] = _safe_float(c(126))
    record[f"{path}.p_dm"] = _map_val(c(127), DM_MAP)
    record[f"{path}.p_dm_mm"] = _safe_float(c(128))
    record[f"{path}.p_ln_metastasis"] = _map_val(c(129), PN_MAP)
    record[f"{path}.p_distant_metastasis"] = _map_val(c(130), PM_META_MAP)
    # col 131: 遠隔転移部位 (テキスト, p_meta_* 将来対応)
    record[f"{path}.p_peritoneal"] = _map_val(c(132), PP_MAP)
    record[f"{path}.p_cytology"] = _parse_cy(c(133))
    record[f"{path}.p_liver"] = _map_val(c(134), PH_MAP)
    record[f"{path}.p_stage"] = _map_val(c(135), P_STAGE_MAP)
    record[f"{path}.p_residual_tumor"] = _parse_residual(c(136))
    record[f"{path}.p_chemo_effect"] = _map_val(c(137), CHEMO_EFFECT_MAP)
    record[f"{path}.p_ln_chemo_effect"] = _map_val(c(138), CHEMO_EFFECT_MAP)

    # Biomarkers
    record[f"{path}.msi_status"] = _map_val(c(139), MSI_MAP)
    record[f"{path}.her2_status"] = _map_val(c(140), HER2_MAP)
    record[f"{path}.pdl1_status"] = _map_val(c(141), PDL1_MAP)
    record[f"{path}.pdl1_cps"] = _safe_float(c(142))
    record[f"{path}.pdl1_tps"] = _safe_float(c(143))

    # ------ lymph_nodes ------
    ln = "lymph_nodes"
    col_idx = 144  # No.1(M) starts at col 144
    for station in ["1", "2", "3a", "3b", "4sa", "4sb", "4d", "5", "6",
                    "7", "8a", "8p", "9", "10", "11p", "11d", "12a",
                    "14v", "16", "19", "20", "108", "110", "111", "112"]:
        m_val = _safe_int(c(col_idx))
        l_val = _safe_int(c(col_idx + 1))
        record[f"{ln}.ln_{station}_m"] = m_val if m_val is not None else 0
        record[f"{ln}.ln_{station}_l"] = l_val if l_val is not None else 0
        col_idx += 2

    # ------ gist_detail ------ (cols 196-202)
    if dc == 4:  # GIST
        gist = "gist_detail"
        record[f"{gist}.gist_kit"] = _map_val(c(196), GIST_IHC_MAP)
        record[f"{gist}.gist_cd34"] = _map_val(c(197), GIST_IHC_MAP)
        record[f"{gist}.gist_desmin"] = _map_val(c(198), GIST_IHC_MAP)
        record[f"{gist}.gist_s100"] = _map_val(c(199), GIST_IHC_MAP)
        record[f"{gist}.gist_mitosis"] = _safe_int(c(200))
        record[f"{gist}.gist_rupture"] = _map_val(c(201), YN_MAP)
        record[f"{gist}.gist_fletcher"] = _map_val(c(202), FLETCHER_MAP)

    # ------ adjuvant_chemo ------ (cols 203-208)
    adj = "adjuvant_chemo"
    adj_yn = c(203)
    if adj_yn and adj_yn != "なし" and adj_yn == "あり":
        record[f"{adj}.adj_yn"] = 1
        record[f"{adj}.adj_start_date"] = _fmt_date(c(204))
        record[f"{adj}.adj_regimen"] = _map_val(c(205), ADJ_REGIMEN_MAP, 99)
        if record[f"{adj}.adj_regimen"] == 99 and c(205):
            record[f"{adj}.adj_regimen_other"] = c(205)
        record[f"{adj}.adj_courses"] = _safe_int(c(206))
        record[f"{adj}.adj_completion"] = _map_val(c(207), CHEMO_COMP_MAP)
        record[f"{adj}.adj_adverse_event"] = c(208)
    else:
        record[f"{adj}.adj_yn"] = 0

    # ------ palliative_chemo ------ (cols 209-229, 1st-5th line)
    pal_yn = c(209)
    if pal_yn and pal_yn != "なし" and pal_yn == "あり":
        for line_n in range(1, 6):
            base_col = 210 + (line_n - 1) * 4  # 210,214,218,222,226
            regimen_text = c(base_col)
            if regimen_text is None:
                continue
            prefix = f"palliative_chemo.line{line_n}"
            record[f"{prefix}_regimen"] = _parse_pal_regimen(regimen_text)
            if record[f"{prefix}_regimen"] == 99 and regimen_text:
                record[f"{prefix}_regimen_other"] = regimen_text
            record[f"{prefix}_start_date"] = _fmt_date(c(base_col + 1))
            record[f"{prefix}_courses"] = _safe_int(c(base_col + 2))
            record[f"{prefix}_adverse_event"] = c(base_col + 3)

    # ------ outcome ------ (cols 10-16)
    out = "outcome"
    rec_yn = _map_val(c(11), RECURRENCE_MAP)
    record[f"{out}.recurrence_yn"] = rec_yn
    record[f"{out}.recurrence_date"] = _fmt_date(c(10))
    record[f"{out}.vital_status"] = _map_val(c(15), VITAL_MAP)
    record[f"{out}.last_alive_date"] = _fmt_date(c(13))
    record[f"{out}.death_date"] = _fmt_date(c(14))
    record[f"{out}.death_cause"] = _map_val(c(16), DEATH_CAUSE_MAP)

    # 再発形式分解
    rec_flags = _decompose_recurrence_sites(c(12))
    for k, v in rec_flags.items():
        record[f"{out}.{k}"] = v

    # ------ None除去 ------
    record = OrderedDict((k, v) for k, v in record.items() if v is not None)

    return record, warnings


# ============================================================
# CSV出力
# ============================================================
def convert_excel_to_csv(excel_path, output_path=None, dry_run=False):
    """Excelファイルを読み込み、UGI_DB互換CSVを出力する。"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    max_row = ws.max_row

    print(f"📖 読み込み: {excel_path}")
    print(f"   シート: {ws.title}, {max_row - 1} 行")

    # 全行変換
    all_records = []
    all_warnings = []
    skipped = 0

    for row in range(2, max_row + 1):
        # 空行チェック (台帳番号 or 患者ID が空ならスキップ)
        if _cell(ws, row, 1) is None and _cell(ws, row, 2) is None:
            skipped += 1
            continue

        record, warnings = convert_row(ws, row)

        # surgery_date 必須チェック
        if "surgery_date" not in record:
            all_warnings.append(f"行{row}: surgery_date が空 → スキップ")
            skipped += 1
            continue

        all_records.append(record)
        for w in warnings:
            all_warnings.append(f"行{row}: {w}")

    print(f"   変換成功: {len(all_records)} 行, スキップ: {skipped} 行")

    # 全レコードの列名を統合
    all_cols = OrderedDict()
    for rec in all_records:
        for k in rec:
            all_cols[k] = True
    col_list = list(all_cols.keys())

    if dry_run:
        print(f"\n🔍 ドライラン: {len(col_list)} カラム")
        for c in col_list[:30]:
            print(f"   {c}")
        if len(col_list) > 30:
            print(f"   ... 他 {len(col_list) - 30} カラム")
        if all_warnings:
            print(f"\n⚠️  警告: {len(all_warnings)} 件")
            for w in all_warnings[:20]:
                print(f"   {w}")
        return

    # CSV出力
    if output_path is None:
        output_path = os.path.splitext(excel_path)[0] + "_ugidb.csv"

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=col_list)
        writer.writeheader()
        for rec in all_records:
            writer.writerow(rec)

    print(f"✅ CSV出力: {output_path}")
    print(f"   {len(all_records)} 行 × {len(col_list)} カラム")

    if all_warnings:
        print(f"\n⚠️  警告: {len(all_warnings)} 件")
        for w in all_warnings[:30]:
            print(f"   {w}")
        if len(all_warnings) > 30:
            print(f"   ... 他 {len(all_warnings) - 30} 件")

    return output_path


# ============================================================
# CLI
# ============================================================
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("使い方: python excel_to_ugidb_csv.py input.xlsx [output.csv] [--dry-run]")
        sys.exit(1)

    excel_path = sys.argv[1]
    output_path = None
    dry_run = "--dry-run" in sys.argv

    for arg in sys.argv[2:]:
        if arg != "--dry-run" and arg.endswith(".csv"):
            output_path = arg

    convert_excel_to_csv(excel_path, output_path, dry_run)
